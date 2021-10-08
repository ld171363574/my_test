# -*- enconding: utf-8 -*-
import psutil
import time
import openpyxl as op


def paprs(times):
    ws = op.Workbook()
    wb = ws.create_sheet(index=0)
    wb.cell(row=1, column=1, value='time')
    wb.cell(row=1, column=2, value='cpu')
    wb.cell(row=1, column=3, value='mem')
    count = 2
    while count < times:
        # 获取当前时间和cpu的占有率
        t = time.localtime()
        cpu_time = '%d:%d:%d' % (t.tm_hour, t.tm_min, t.tm_sec)
        cpu_res = psutil.cpu_percent()
        mem = psutil.virtual_memory()
        mems = float(mem.used / 1024 / 1024)

        # 1秒采集一次
        time.sleep(1)
        # 写入数据
        wb.cell(row=count, column=1, value=cpu_time)
        wb.cell(row=count, column=2, value=cpu_res)
        wb.cell(row=count, column=3, value=mems)
        count += 1
        ws.save('cpu_men.xlsx')


if __name__ == "__main__":
    paprs(times=30)
